import telebot
from telebot import types
import sqlite3
from datetime import datetime, date, timedelta
import logging
import traceback
import time
import requests
import telebot.apihelper
from PIL import Image
import pytesseract
import io
from difflib import SequenceMatcher
import os
import sys

# ==================== LOGGING ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==================== SQLITE ADAPTER ====================
def adapt_datetime(dt):
    return dt.isoformat()

def adapt_date(dt):
    return dt.isoformat()

def convert_datetime(s):
    try:
        return datetime.fromisoformat(s.decode('utf-8')) if s else None
    except (ValueError, AttributeError):
        return None

def convert_date(s):
    try:
        return date.fromisoformat(s.decode('utf-8')) if s else None
    except (ValueError, AttributeError):
        return None

sqlite3.register_adapter(datetime, adapt_datetime)
sqlite3.register_adapter(date, adapt_date)
sqlite3.register_converter("TIMESTAMP", convert_datetime)
sqlite3.register_converter("DATE", convert_date)

# ==================== KONFIGURATSIYA ====================
BOT_TOKEN = '8523430941:AAE3lzJ3aFF2ss6JJ3S5e_ERgAj29DhJ0VM'
CHANNEL_ID = '-1003543686638'

# Doimiy adminlar (kod ichida - o'zgarmas)
ADMIN_IDS = [
    8517530604,  # Asosiy admin
]

DB_NAME = 'students.db'
GROQ_API_KEY = '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'
GROQ_API_URL = 'https://api.groq.com/openai/v1/chat/completions'

# ==================== GLOBAL ====================
active_contest = None
user_states = {}
bot = None

# ==================== XAVFSIZ FUNKSIYALAR ====================
def safe_execute(func, *args, max_retries=3, default_return=None, **kwargs):
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except telebot.apihelper.ApiTelegramException as e:
            error_msg = str(e).lower()
            if any(x in error_msg for x in ['query is too old', 'timeout', 'invalid', 'blocked', 'chat not found']):
                logger.warning(f"⚠️ Ignored Telegram error: {e}")
                return default_return
            if attempt == max_retries - 1:
                logger.error(f"❌ {func.__name__} failed: {e}")
                return default_return
            time.sleep(2 ** attempt)
        except (ConnectionError, ConnectionResetError, requests.exceptions.ConnectionError) as e:
            logger.warning(f"⚠️ Connection error: {e}")
            if attempt == max_retries - 1:
                return default_return
            time.sleep(2 ** attempt)
        except Exception as e:
            logger.error(f"❌ {func.__name__} error: {e}")
            logger.error(traceback.format_exc())
            if attempt == max_retries - 1:
                return default_return
            time.sleep(1)
    return default_return

def safe_send_message(chat_id, text, **kwargs):
    return safe_execute(bot.send_message, chat_id, text, **kwargs)

def safe_send_document(chat_id, document, **kwargs):
    return safe_execute(bot.send_document, chat_id, document, **kwargs)

def safe_send_photo(chat_id, photo, **kwargs):
    return safe_execute(bot.send_photo, chat_id, photo, **kwargs)

def safe_edit_message_text(text, chat_id, message_id, **kwargs):
    return safe_execute(bot.edit_message_text, text, chat_id, message_id, **kwargs)

def safe_answer_callback_query(callback_query_id, text="", show_alert=False, **kwargs):
    return safe_execute(bot.answer_callback_query, callback_query_id, text=text, show_alert=show_alert, **kwargs)

def escape_html(text):
    if text is None:
        return ""
    text = str(text)
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace('"', '&quot;')
    text = text.replace("'", "&#x27;")
    return text

# ==================== DATABASE ====================
def get_db_connection():
    try:
        conn = sqlite3.connect(
            DB_NAME,
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES,
            timeout=10,
            check_same_thread=False
        )
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        logger.error(f"❌ DB connection error: {e}")
        return None

def safe_db_execute(query, params=(), fetch_one=False, fetch_all=False, commit=False):
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            return None
        cursor = conn.cursor()
        cursor.execute(query, params)
        result = None
        if fetch_one:
            result = cursor.fetchone()
        elif fetch_all:
            result = cursor.fetchall()
        elif commit:
            conn.commit()
            result = cursor.lastrowid if 'INSERT' in query.upper() else True
        return result
    except sqlite3.OperationalError as e:
        logger.error(f"❌ SQL error: {e} | Query: {query[:100]}")
        return None
    except Exception as e:
        logger.error(f"❌ DB error: {e}")
        return None
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

# ==================== BUG FIX #1: is_admin safe_db_execute dan KEYIN ====================
# is_admin() funksiyasi safe_db_execute() dan KEYIN aniqlanmoqda
def is_admin(user_id):
    """
    Foydalanuvchi admin ekanligini tekshirish.
    Avval ADMIN_IDS (kod ichida), keyin DB dagi admins jadvalini tekshiradi.
    """
    if user_id in ADMIN_IDS:
        return True
    # DB dan tekshirish
    result = safe_db_execute(
        'SELECT user_id FROM admins WHERE user_id = ? AND is_active = 1',
        (user_id,),
        fetch_one=True
    )
    return result is not None

# ==================== BUG FIX #2: _save_admin GLOBAL funksiya ====================
def save_admin_to_db(new_admin_id, full_name, username, added_by):
    """
    Yangi adminni DB ga saqlash - GLOBAL funksiya (scope xatosidan xoli).
    """
    existing = safe_db_execute(
        'SELECT user_id FROM admins WHERE user_id = ?',
        (new_admin_id,),
        fetch_one=True
    )
    if existing:
        safe_db_execute(
            'UPDATE admins SET is_active = 1, full_name = ?, username = ?, added_by = ?, added_at = ? WHERE user_id = ?',
            (full_name, username, added_by, datetime.now(), new_admin_id),
            commit=True
        )
    else:
        safe_db_execute(
            'INSERT INTO admins (user_id, full_name, username, added_by, added_at) VALUES (?, ?, ?, ?, ?)',
            (new_admin_id, full_name, username, added_by, datetime.now()),
            commit=True
        )
    # Students jadvaliga ham qo'shish (agar yo'q bo'lsa)
    safe_db_execute(
        'INSERT OR IGNORE INTO students (user_id, full_name, username, registered_at) VALUES (?, ?, ?, ?)',
        (new_admin_id, full_name, username, datetime.now()),
        commit=True
    )
    logger.info(f"✅ Admin saqlandi: {new_admin_id} ({full_name})")

def init_db():
    try:
        conn = get_db_connection()
        if not conn:
            return False
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                user_id INTEGER PRIMARY KEY,
                full_name TEXT NOT NULL,
                username TEXT,
                registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active INTEGER DEFAULT 1
            )
        ''')

        # Adminlar jadvali
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS admins (
                user_id INTEGER PRIMARY KEY,
                full_name TEXT NOT NULL,
                username TEXT,
                added_by INTEGER NOT NULL,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active INTEGER DEFAULT 1
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                homework_text TEXT NOT NULL,
                assignment_date DATE NOT NULL DEFAULT (date('now')),
                sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active INTEGER DEFAULT 1
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                full_name TEXT NOT NULL,
                homework_text TEXT,
                homework_file TEXT,
                file_type TEXT,
                assignment_id INTEGER,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'pending',
                reviewed_at TIMESTAMP,
                reviewer_id INTEGER,
                ai_check_result TEXT,
                ai_checked_at TIMESTAMP,
                rejection_reason TEXT,
                FOREIGN KEY (user_id) REFERENCES students (user_id),
                FOREIGN KEY (assignment_id) REFERENCES assignments (id)
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS statistics (
                date TEXT PRIMARY KEY,
                total_submissions INTEGER DEFAULT 0,
                approved_submissions INTEGER DEFAULT 0,
                rejected_submissions INTEGER DEFAULT 0
            )
        ''')

        # Contests: media_file_id va media_type qo'shildi (to'g'ri javob kanalga chiqmaydi)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS contests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                problem_text TEXT NOT NULL,
                correct_answer TEXT NOT NULL,
                deadline TIMESTAMP NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active INTEGER DEFAULT 1,
                media_file_id TEXT,
                media_type TEXT
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS contest_submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contest_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                full_name TEXT NOT NULL,
                answer TEXT NOT NULL,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_correct INTEGER DEFAULT 0,
                rank_position INTEGER,
                FOREIGN KEY (contest_id) REFERENCES contests(id),
                FOREIGN KEY (user_id) REFERENCES students(user_id)
            )
        ''')

        # Eski DB ga ustunlar qo'shish (migrate)
        for col_sql in [
            'ALTER TABLE contests ADD COLUMN media_file_id TEXT',
            'ALTER TABLE contests ADD COLUMN media_type TEXT',
        ]:
            try:
                cursor.execute(col_sql)
            except:
                pass

        conn.commit()
        conn.close()
        logger.info("✅ Database tayyor")
        return True
    except Exception as e:
        logger.error(f"❌ Database init error: {e}")
        return False

# ==================== YORDAMCHI FUNKSIYALAR ====================
def clear_user_state(user_id):
    if user_id in user_states:
        del user_states[user_id]

def is_registered(user_id):
    result = safe_db_execute(
        'SELECT user_id FROM students WHERE user_id = ? AND is_active = 1',
        (user_id,),
        fetch_one=True
    )
    return result is not None

def get_student_info(user_id):
    result = safe_db_execute(
        'SELECT * FROM students WHERE user_id = ?',
        (user_id,),
        fetch_one=True
    )
    return dict(result) if result else None

def get_all_students():
    results = safe_db_execute(
        'SELECT * FROM students WHERE is_active = 1 ORDER BY registered_at DESC',
        fetch_all=True
    )
    return [dict(row) for row in results] if results else []

def get_all_admins():
    results = safe_db_execute(
        'SELECT * FROM admins WHERE is_active = 1 ORDER BY added_at DESC',
        fetch_all=True
    )
    return [dict(row) for row in results] if results else []

def get_current_assignment():
    today = date.today().strftime('%Y-%m-%d')
    result = safe_db_execute(
        'SELECT * FROM assignments WHERE assignment_date = ? AND is_active = 1 ORDER BY sent_at DESC LIMIT 1',
        (today,),
        fetch_one=True
    )
    return result

def get_retry_count(user_id, assignment_id):
    result = safe_db_execute(
        'SELECT COUNT(*) as count FROM submissions WHERE user_id = ? AND assignment_id = ?',
        (user_id, assignment_id),
        fetch_one=True
    )
    return result['count'] if result else 0

def update_statistics(status):
    today = datetime.now().strftime('%Y-%m-%d')
    exists = safe_db_execute(
        'SELECT date FROM statistics WHERE date = ?',
        (today,),
        fetch_one=True
    )
    if not exists:
        safe_db_execute(
            'INSERT INTO statistics (date, total_submissions, approved_submissions, rejected_submissions) VALUES (?, 0, 0, 0)',
            (today,),
            commit=True
        )
    if status == 'approved':
        safe_db_execute(
            'UPDATE statistics SET approved_submissions = approved_submissions + 1 WHERE date = ?',
            (today,),
            commit=True
        )
    elif status == 'rejected':
        safe_db_execute(
            'UPDATE statistics SET rejected_submissions = rejected_submissions + 1 WHERE date = ?',
            (today,),
            commit=True
        )

# ==================== KLAVIATURALAR ====================
def get_main_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton('📝 Uyga vazifa topshirish'),
        types.KeyboardButton('📊 Statistika'),
        types.KeyboardButton('✍️ Javob yuborish'),
        types.KeyboardButton('🏆 Reyting'),
        types.KeyboardButton('❓ Yordam')
    )
    return markup

def get_admin_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton('📤 Uyga vazifa yuborish'),
        types.KeyboardButton('📊 Statistika'),
        types.KeyboardButton('🏆 IT Misol'),
        types.KeyboardButton('🏆 Reyting'),
        types.KeyboardButton('📥 Excel'),
        types.KeyboardButton('👨‍💼 Admin panel'),
        types.KeyboardButton('➕ Admin boshqaruv'),
        types.KeyboardButton('❓ Yordam')
    )
    return markup

# ==================== OCR ====================
def extract_text_from_image(file_id):
    try:
        file_info = bot.get_file(file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        image = Image.open(io.BytesIO(downloaded_file))
        text = pytesseract.image_to_string(image, lang='uzb+eng+rus')
        logger.info(f"✅ OCR: {len(text)} belgi")
        return text.strip()
    except Exception as e:
        logger.error(f"❌ OCR error: {e}")
        return None

# ==================== AI TEKSHIRUV ====================
def check_homework_with_ai(homework_text, student_name, submission_id):
    """
    AI orqali uyga vazifani tekshirish.
    Groq Llama 3.3 70B ishlatiladi.
    """
    try:
        assignment = get_current_assignment()
        if not assignment:
            return {'status': 'error', 'message': "❌ Joriy vazifa topilmadi"}

        assignment_text = assignment[1]

        # O'xshash vazifalarni tekshirish
        duplicates = safe_db_execute(
            'SELECT full_name FROM submissions WHERE homework_text = ? AND full_name != ? AND status != "rejected" LIMIT 5',
            (homework_text, student_name),
            fetch_all=True
        )

        duplicate_msg = ""
        if duplicates:
            duplicate_msg = "⚠️ DIQQAT! O'xshash vazifalar topildi:\n"
            for dup in duplicates:
                duplicate_msg += f"• {escape_html(dup['full_name'])}\n"
            duplicate_msg += "\n"

        prompt = f"""Siz professional o'qituvchisiz. O'quvchi javobini quyidagi standart vazifaga nisbatan tekshiring:

STANDART VAZIFA:
{assignment_text}

O'QUVCHI: {student_name}
JAVOB:
{homework_text}

Quyidagilarni baholang:
1. Grammatika va imlo (xatolarni ko'rsating)
2. Mazmun to'g'riligi (standart vazifaga mos kelishi, 0-100%)
3. Tuzilish va tartib
4. Umumiy baho

QOIDA: Agar to'g'rilik 70% dan past bo'lsa - RAD ETING.

JAVOB FORMATI (HTML yo'q, faqat oddiy matn):
✅/❌ BAHO: [Qabul qilindi / Rad etildi]
📊 TO'G'RILIK: [0-100]%
📝 TAHLIL:
- Grammatika: [xatolar yoki "Xato yo'q"]
- Mazmun: [baholash]
- Tuzilish: [baholash]
💡 TAVSIYA: [qisqa maslahat]"""

        headers = {
            'Authorization': f'Bearer {GROQ_API_KEY}',
            'Content-Type': 'application/json'
        }
        data = {
            'model': 'llama-3.3-70b-versatile',
            'messages': [
                {
                    'role': 'system',
                    'content': "Siz qattiq va adolatli o'qituvchisiz. Faqat standart formatda javob bering. HTML ishlatmang."
                },
                {'role': 'user', 'content': prompt}
            ],
            'temperature': 0.5,
            'max_tokens': 800
        }

        response = requests.post(GROQ_API_URL, headers=headers, json=data, timeout=30)

        if response.status_code == 200:
            result = response.json()
            ai_response = result['choices'][0]['message']['content']

            # Natijani DB ga saqlash
            safe_db_execute(
                'UPDATE submissions SET ai_check_result = ?, ai_checked_at = ? WHERE id = ?',
                (ai_response, datetime.now(), submission_id),
                commit=True
            )

            full_response = f"{duplicate_msg}{ai_response}\n\n🤖 <i>AI: Groq Llama 3.3 70B</i>"
            return {'status': 'success', 'message': full_response}
        else:
            logger.error(f"❌ Groq API error: {response.status_code} - {response.text[:200]}")
            return {'status': 'error', 'message': f"❌ AI xatolik: {response.status_code}"}

    except Exception as e:
        logger.error(f"❌ AI check error: {e}")
        logger.error(traceback.format_exc())
        return {'status': 'error', 'message': f"❌ AI xatolik: {str(e)}"}

# ==================== BROADCAST ====================
def broadcast_assignment(assignment_text, assignment_id, assignment_date):
    students = get_all_students()
    sent_count = 0
    for student in students:
        if not is_admin(student['user_id']):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('📝 Ushbu vazifani topshirish'))
            success = safe_send_message(
                student['user_id'],
                f"📚 <b>Yangi uyga vazifa! ({assignment_date})</b>\n\n"
                f"{escape_html(assignment_text)}\n\n"
                f"🔢 ID: #{assignment_id}\n\n"
                f"📝 'Ushbu vazifani topshirish' tugmasini bosing.",
                parse_mode='HTML',
                reply_markup=markup
            )
            if success:
                sent_count += 1
    logger.info(f"✅ Broadcast: {sent_count} ta o'quvchi")
    return sent_count

# ==================== HANDLERLARNI RO'YXATGA OLISH ====================
def register_handlers(bot_instance):
    global bot
    bot = bot_instance

    # ==================== /start va /cancel ====================
    @bot.message_handler(commands=['start', 'cancel', 'menu'])
    def start(message):
        user_id = message.from_user.id
        first_name = message.from_user.first_name or "Noma'lum"
        last_name = message.from_user.last_name or ""
        username = message.from_user.username
        logger.info(f"▶️ /{message.text.split()[0].replace('/','')} - User: {user_id}")

        # MUHIM: Har doim eski holatni tozalash
        # Bu /start, /cancel, /menu bosilganda qolib ketgan state ni o'chiradi
        if user_id in user_states:
            logger.info(f"🧹 State tozalandi: {user_id} ({user_states[user_id]})")
            clear_user_state(user_id)

        if is_admin(user_id):
            if not is_registered(user_id):
                full_name = f"{first_name} {last_name}".strip()
                safe_db_execute(
                    'INSERT OR REPLACE INTO students (user_id, full_name, username, registered_at) VALUES (?, ?, ?, ?)',
                    (user_id, full_name, username, datetime.now()),
                    commit=True
                )
            student = get_student_info(user_id)
            safe_send_message(
                message.chat.id,
                f"🎓 Salom, Admin {escape_html(student['full_name'])}! 👨‍💼\n\nBoshqaruv paneli:",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
            return

        student = get_student_info(user_id)
        if student:
            safe_send_message(
                message.chat.id,
                f"🎓 Salom, {escape_html(student['full_name'])}!\n\nQuyidagi bo'limlardan birini tanlang:",
                parse_mode='HTML',
                reply_markup=get_main_keyboard()
            )
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton("✅ Ro'yxatdan o'tish"))
            safe_send_message(
                message.chat.id,
                "👋 Assalomu alaykum!\n\nBotdan foydalanish uchun ro'yxatdan o'ting.\n\n"
                "📝 Ism va familiyangizni to'liq kiriting:\n(Masalan: Muhammad Aliyev)",
                reply_markup=markup
            )

    # ==================== YORDAM ====================
    @bot.message_handler(func=lambda m: m.text == '❓ Yordam')
    def help_command(message):
        user_id = message.from_user.id
        if is_admin(user_id):
            text = (
                "👨‍💼 <b>Admin yordam</b>\n\n"
                "📤 <b>Uyga vazifa yuborish</b> — barcha o'quvchilarga vazifa\n"
                "🏆 <b>IT Misol</b> — musobaqa (rasm/video bilan)\n"
                "➕ <b>Admin boshqaruv</b> — admin qo'shish/o'chirish\n"
                "📊 <b>Statistika</b> — hisobotlar\n"
                "📥 <b>Excel</b> — ma'lumotlarni yuklab olish\n"
                "👨‍💼 <b>Admin panel</b> — o'quvchilar, tozalash\n"
                "🤖 <b>AI tekshiruv</b> — topshiriqlarda AI tugmasi bilan"
            )
        else:
            text = (
                "📚 <b>Bot haqida</b>\n\n"
                "1️⃣ Ro'yxatdan o'ting\n"
                "2️⃣ Vazifani topshiring (matn/rasm/fayl)\n"
                "3️⃣ AI foiz bilan tekshiradi\n"
                "4️⃣ O'qituvchi tasdiqlaydi\n\n"
                "🏆 <b>IT Misol</b> — tez yechish musobaqasi\n"
                "📊 <b>Statistika</b> — shaxsiy natijalar"
            )
        safe_send_message(message.chat.id, text, parse_mode='HTML')

    # ==================== RO'YXATDAN O'TISH ====================
    @bot.message_handler(func=lambda m: m.text in ["✅ Ro'yxatdan o'tish", "Ro'yxatdan o'tish"])
    def register_start(message):
        user_id = message.from_user.id
        if is_admin(user_id) or is_registered(user_id):
            safe_send_message(message.chat.id, "✅ Siz allaqachon ro'yxatdan o'tgansiz!")
            return
        user_states[user_id] = 'registering_name'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('❌ Bekor qilish'))
        safe_send_message(
            message.chat.id,
            "✍️ Ism va familiyangizni to'liq kiriting:\n(Masalan: Muhammad Aliyev, kamida 10 harf)",
            reply_markup=markup
        )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id] == 'registering_name'
    )
    def register_complete(message):
        user_id = message.from_user.id
        full_name = message.text.strip()
        username = message.from_user.username

        if full_name == '❌ Bekor qilish':
            clear_user_state(user_id)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton("✅ Ro'yxatdan o'tish"))
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=markup)
            return

        if len(full_name) < 10:
            safe_send_message(message.chat.id, "❌ Juda qisqa. Kamida 10 harf kiriting.")
            return

        safe_db_execute(
            'INSERT OR REPLACE INTO students (user_id, full_name, username, registered_at) VALUES (?, ?, ?, ?)',
            (user_id, full_name, username, datetime.now()),
            commit=True
        )
        clear_user_state(user_id)
        safe_send_message(
            message.chat.id,
            f"✅ Muvaffaqiyatli ro'yxatdan o'tdingiz!\n\n"
            f"👤 Ism: {escape_html(full_name)}\n"
            f"📅 Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
            f"Uyga vazifalarni topshirishingiz mumkin! 📚",
            parse_mode='HTML',
            reply_markup=get_main_keyboard()
        )
        logger.info(f"✅ Ro'yxat: {full_name} (ID: {user_id})")

    # ==================== STATISTIKA ====================
    @bot.message_handler(func=lambda m: m.text == '📊 Statistika')
    def show_statistics(message):
        user_id = message.from_user.id
        if not is_registered(user_id):
            safe_send_message(message.chat.id, "❌ Avval ro'yxatdan o'ting!")
            return

        student = get_student_info(user_id)

        total = (safe_db_execute(
            'SELECT COUNT(*) as c FROM submissions WHERE user_id = ?', (user_id,), fetch_one=True
        ) or {}).get('c', 0)
        approved = (safe_db_execute(
            'SELECT COUNT(*) as c FROM submissions WHERE user_id = ? AND status = "approved"', (user_id,), fetch_one=True
        ) or {}).get('c', 0)
        rejected = (safe_db_execute(
            'SELECT COUNT(*) as c FROM submissions WHERE user_id = ? AND status = "rejected"', (user_id,), fetch_one=True
        ) or {}).get('c', 0)

        today = datetime.now().strftime('%Y-%m-%d')
        today_stats = safe_db_execute(
            'SELECT total_submissions, approved_submissions, rejected_submissions FROM statistics WHERE date = ?',
            (today,), fetch_one=True
        )

        text = f"📊 <b>{escape_html(student['full_name'])} - Statistika</b>\n\n"
        text += f"📈 Umumiy: {total}\n✅ Tasdiqlangan: {approved}\n❌ Rad etilgan: {rejected}\n\n"

        if today_stats:
            text += (
                f"📅 Bugun:\n"
                f"• Topshirilgan: {today_stats['total_submissions'] or 0}\n"
                f"• Tasdiqlangan: {today_stats['approved_submissions'] or 0}\n"
                f"• Rad etilgan: {today_stats['rejected_submissions'] or 0}\n"
            )
        else:
            text += "📅 Bugun: Hech narsa yo'q\n"

        if total > 0:
            text += f"\n📈 Muvaffaqiyat: {(approved / total * 100):.1f}%"

        safe_send_message(message.chat.id, text, parse_mode='HTML')

    # ==================== UYGA VAZIFA YUBORISH (ADMIN) ====================
    @bot.message_handler(
        func=lambda m: m.text == '📤 Uyga vazifa yuborish' and is_admin(m.from_user.id)
    )
    def broadcast_start(message):
        user_id = message.from_user.id
        if user_id in user_states:
            logger.info(f"🧹 Broadcast boshlashdan oldin state tozalandi: {user_states[user_id]}")
            clear_user_state(user_id)
        user_states[user_id] = 'broadcasting_homework'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('❌ Bekor qilish'))
        safe_send_message(
            message.chat.id,
            "📤 Uyga vazifani kiriting (matn):\n\nYuborganingizdan so'ng barcha o'quvchilarga jo'natiladi.",
            reply_markup=markup
        )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id] == 'broadcasting_homework'
    )
    def broadcast_complete(message):
        user_id = message.from_user.id
        assignment_text = message.text.strip()

        if assignment_text == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if len(assignment_text) < 10:
            safe_send_message(message.chat.id, "❌ Juda qisqa. Batafsilroq kiriting.")
            return

        today = date.today().strftime('%Y-%m-%d')
        assignment_id = safe_db_execute(
            'INSERT INTO assignments (homework_text, assignment_date) VALUES (?, ?)',
            (assignment_text, today), commit=True
        )
        if not assignment_id:
            safe_send_message(message.chat.id, "❌ Saqlashda xatolik!")
            return

        safe_db_execute(
            'UPDATE assignments SET is_active = 0 WHERE assignment_date = ? AND id != ?',
            (today, assignment_id), commit=True
        )

        sent_count = broadcast_assignment(assignment_text, assignment_id, today)
        clear_user_state(user_id)
        safe_send_message(
            message.chat.id,
            f"✅ Broadcast muvaffaqiyatli!\n\n"
            f"📝 Vazifa: {escape_html(assignment_text[:50])}...\n"
            f"🔢 ID: #{assignment_id}\n📅 {today}\n"
            f"📢 Yuborildi: {sent_count} ta o'quvchi",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
        safe_send_message(
            CHANNEL_ID,
            f"📚 <b>Yangi vazifa ({today})</b>\n\n"
            f"📝 {escape_html(assignment_text)}\n"
            f"🔢 ID: #{assignment_id}\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            parse_mode='HTML'
        )

    # ==================== UYGA VAZIFA TOPSHIRISH (O'QUVCHI) ====================
    @bot.message_handler(
        func=lambda m: m.text in ['📝 Ushbu vazifani topshirish', '📝 Uyga vazifa topshirish']
    )
    def submit_homework_start(message):
        user_id = message.from_user.id
        # Oldingi state ni tozalash (masalan, oldin boshqa jarayon boshlangan bo'lsa)
        if user_id in user_states:
            logger.info(f"🧹 Homework topshirishdan oldin state tozalandi: {user_states[user_id]}")
            clear_user_state(user_id)
        if not is_registered(user_id):
            safe_send_message(message.chat.id, "❌ Avval ro'yxatdan o'ting!")
            return

        current_assignment = get_current_assignment()
        if not current_assignment:
            safe_send_message(message.chat.id, "❌ Hozircha yangi vazifa yo'q!")
            return

        retry_count = get_retry_count(user_id, current_assignment[0])
        if retry_count >= 3:
            safe_send_message(
                message.chat.id,
                f"❌ Siz bu vazifani {retry_count}/3 marta topshirdingiz. Limit tugadi!"
            )
            return

        user_states[user_id] = 'submitting_homework'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('❌ Bekor qilish'))
        retry_msg = f"\n🔄 Urinish: {retry_count + 1}/3" if retry_count > 0 else ""
        safe_send_message(
            message.chat.id,
            f"📤 Joriy vazifa ({current_assignment[2]}):\n\n"
            f"{escape_html(current_assignment[1])}\n\n"
            f"📝 Vazifangizni yuboring:\n"
            f"• 📄 Matn  • 🖼 Rasm (OCR)  • 📎 Fayl  • 🎥 Video  • 🎵 Audio\n\n"
            f"🤖 AI foiz bilan tekshiradi!{retry_msg}\n"
            f"🔢 Vazifa ID: #{current_assignment[0]}",
            parse_mode='HTML',
            reply_markup=markup
        )

    @bot.message_handler(
        content_types=['text', 'document', 'photo', 'video', 'audio', 'voice'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id] == 'submitting_homework'
    )
    def receive_homework(message):
        user_id = message.from_user.id
        if not is_registered(user_id):
            safe_send_message(message.chat.id, "❌ Avval ro'yxatdan o'ting!")
            return

        # BUG FIX #3: message.text None bo'lganda .strip() xatosi oldini olish
        if message.text and message.text.strip() == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_main_keyboard())
            return

        student = get_student_info(user_id)
        current_assignment = get_current_assignment()
        if not current_assignment:
            safe_send_message(message.chat.id, "❌ Joriy vazifa topilmadi!")
            return

        assignment_id = current_assignment[0]
        retry_count = get_retry_count(user_id, assignment_id)
        if retry_count >= 3:
            safe_send_message(message.chat.id, "❌ Limit tugadi (3/3)!")
            return

        homework_text = None
        homework_file = None
        file_type = None

        if message.document:
            homework_file = message.document.file_id
            file_type = 'document'
            homework_text = message.document.file_name or "📎 Fayl yuborildi"
        elif message.photo:
            homework_file = message.photo[-1].file_id
            file_type = 'photo'
            extracted = extract_text_from_image(homework_file)
            homework_text = extracted if extracted else "🖼 Rasm yuborildi (OCR muvaffaqiyatsiz)"
        elif message.video:
            homework_file = message.video.file_id
            file_type = 'video'
            homework_text = "🎥 Video yuborildi"
        elif message.audio:
            homework_file = message.audio.file_id
            file_type = 'audio'
            homework_text = "🎵 Audio yuborildi"
        elif message.voice:
            homework_file = message.voice.file_id
            file_type = 'voice'
            homework_text = "🎤 Ovozli xabar yuborildi"
        elif message.text:
            homework_text = message.text.strip()
        
        if not homework_text or not homework_text.strip():
            safe_send_message(message.chat.id, "❌ Hech narsa yuborilmadi!")
            return

        submission_id = safe_db_execute(
            'INSERT INTO submissions (user_id, full_name, homework_text, homework_file, file_type, assignment_id, submitted_at) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (user_id, student['full_name'], homework_text, homework_file, file_type, assignment_id, datetime.now()),
            commit=True
        )
        if not submission_id:
            safe_send_message(message.chat.id, "❌ Saqlashda xatolik!")
            return

        today = datetime.now().strftime('%Y-%m-%d')
        safe_db_execute(
            'INSERT OR IGNORE INTO statistics (date, total_submissions, approved_submissions, rejected_submissions) VALUES (?, 0, 0, 0)',
            (today,), commit=True
        )
        safe_db_execute(
            'UPDATE statistics SET total_submissions = total_submissions + 1 WHERE date = ?',
            (today,), commit=True
        )

        retry_msg = f"\n🔄 Urinish: {retry_count + 1}/3" if retry_count > 0 else ""
        post_text = (
            f"📚 <b>Yangi topshiriq #{submission_id}</b>\n"
            f"👤 {escape_html(student['full_name'])}\n"
            f"🆔 <code>{user_id}</code>\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
            f"📝 Vazifa #{assignment_id}:\n"
            f"{escape_html(homework_text[:500])}{'...' if len(homework_text) > 500 else ''}"
            f"{retry_msg}"
        )

        # Inline tugmalar (AI tekshiruv tugmasi bor)
        markup = types.InlineKeyboardMarkup(row_width=2)
        btn_approve = types.InlineKeyboardButton('✅ Tasdiqlash', callback_data=f'approve_{submission_id}')
        btn_reject = types.InlineKeyboardButton('❌ Rad etish', callback_data=f'reject_with_reason_{submission_id}')
        btn_ai = types.InlineKeyboardButton('🤖 AI tekshirish', callback_data=f'ai_check_{submission_id}')
        markup.add(btn_approve, btn_reject)

        if file_type == 'photo':
            btn_ai_ocr = types.InlineKeyboardButton('🤖 OCR+AI', callback_data=f'ai_check_ocr_{submission_id}')
            markup.add(btn_ai, btn_ai_ocr)
        else:
            markup.add(btn_ai)

        # Kanalga yuborish
        if homework_file:
            if file_type == 'document':
                safe_send_document(CHANNEL_ID, homework_file, caption=post_text, reply_markup=markup, parse_mode='HTML')
            elif file_type == 'photo':
                safe_send_photo(CHANNEL_ID, homework_file, caption=post_text, reply_markup=markup, parse_mode='HTML')
            elif file_type == 'video':
                safe_execute(bot.send_video, CHANNEL_ID, homework_file, caption=post_text, reply_markup=markup, parse_mode='HTML')
            elif file_type == 'audio':
                safe_execute(bot.send_audio, CHANNEL_ID, homework_file, caption=post_text, reply_markup=markup, parse_mode='HTML')
            elif file_type == 'voice':
                safe_execute(bot.send_voice, CHANNEL_ID, homework_file, caption=post_text, reply_markup=markup, parse_mode='HTML')
        else:
            safe_send_message(CHANNEL_ID, post_text, reply_markup=markup, parse_mode='HTML')

        clear_user_state(user_id)
        safe_send_message(
            message.chat.id,
            f"✅ Vazifangiz yuborildi!\n"
            f"🤖 AI tugmasi orqali tekshirilishi mumkin.\n"
            f"📢 O'qituvchi ko'rib chiqadi.{retry_msg}\n"
            f"🔢 Topshiriq ID: #{submission_id}",
            reply_markup=get_main_keyboard()
        )

    # ==================== INLINE: TASDIQLASH / RAD / AI ====================
    @bot.callback_query_handler(
        func=lambda call: call.data.startswith(('approve_', 'reject_with_reason_', 'ai_check_'))
    )
    def handle_inline_buttons(call):
        user_id = call.from_user.id
        if not is_admin(user_id):
            safe_answer_callback_query(call.id, "❌ Ruxsat yo'q!", show_alert=True)
            return

        data = call.data
        submission_id = int(data.split('_')[-1])
        submission = safe_db_execute(
            'SELECT * FROM submissions WHERE id = ?', (submission_id,), fetch_one=True
        )

        if not submission:
            safe_answer_callback_query(call.id, "❌ Topilmadi!", show_alert=True)
            return
        if submission['status'] != 'pending':
            safe_answer_callback_query(call.id, f"❌ Allaqachon: {submission['status']}", show_alert=True)
            return

        # --- AI tekshiruv (OCR + AI) ---
        if data.startswith('ai_check_ocr_'):
            if submission['file_type'] != 'photo':
                safe_answer_callback_query(call.id, "❌ Faqat rasm uchun!", show_alert=True)
                return
            safe_answer_callback_query(call.id, "🤖 OCR + AI ishlamoqda...")
            extracted = extract_text_from_image(submission['homework_file'])
            if extracted:
                ai_result = check_homework_with_ai(extracted, submission['full_name'], submission_id)
                if ai_result['status'] == 'success':
                    new_text = (
                        f"{call.message.text or call.message.caption or ''}\n\n"
                        f"📄 <b>OCR matni:</b>\n{escape_html(extracted[:300])}...\n\n"
                        f"{ai_result['message']}"
                    )
                    safe_execute(
                        bot.edit_message_caption,
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        caption=new_text[:1024],
                        parse_mode='HTML'
                    )
                else:
                    safe_send_message(call.message.chat.id, ai_result['message'])
            else:
                safe_send_message(call.message.chat.id, "❌ OCR muvaffaqiyatsiz!")

            # Tugmalarni yangilash: faqat Tasdiqlash/Rad qoladi
            new_markup = types.InlineKeyboardMarkup(row_width=2)
            new_markup.add(
                types.InlineKeyboardButton('✅ Tasdiqlash', callback_data=f'approve_{submission_id}'),
                types.InlineKeyboardButton('❌ Rad etish', callback_data=f'reject_with_reason_{submission_id}')
            )
            safe_execute(bot.edit_message_reply_markup,
                         call.message.chat.id, call.message.message_id, reply_markup=new_markup)

        # --- AI tekshiruv (oddiy matn) ---
        elif data.startswith('ai_check_'):
            safe_answer_callback_query(call.id, "🤖 AI tekshiruv boshlanmoqda...")
            ai_result = check_homework_with_ai(
                submission['homework_text'], submission['full_name'], submission_id
            )
            if ai_result['status'] == 'success':
                # Rasm/video caption bo'lsa - caption yangilash, aks holda matn
                if call.message.content_type in ('photo', 'video', 'document', 'audio', 'voice'):
                    new_caption = (
                        f"{call.message.caption or ''}\n\n{ai_result['message']}"
                    )
                    safe_execute(
                        bot.edit_message_caption,
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        caption=new_caption[:1024],
                        parse_mode='HTML'
                    )
                else:
                    new_text = f"{call.message.text or ''}\n\n{ai_result['message']}"
                    safe_edit_message_text(
                        new_text, call.message.chat.id, call.message.message_id,
                        parse_mode='HTML', disable_web_page_preview=True
                    )
            else:
                safe_send_message(call.message.chat.id, ai_result['message'])

            new_markup = types.InlineKeyboardMarkup(row_width=2)
            new_markup.add(
                types.InlineKeyboardButton('✅ Tasdiqlash', callback_data=f'approve_{submission_id}'),
                types.InlineKeyboardButton('❌ Rad etish', callback_data=f'reject_with_reason_{submission_id}')
            )
            safe_execute(bot.edit_message_reply_markup,
                         call.message.chat.id, call.message.message_id, reply_markup=new_markup)

        # --- Tasdiqlash ---
        elif data.startswith('approve_'):
            safe_db_execute(
                'UPDATE submissions SET status = "approved", reviewed_at = ?, reviewer_id = ? WHERE id = ?',
                (datetime.now(), user_id, submission_id), commit=True
            )
            update_statistics('approved')

            # Xabarni yangilash
            approve_text = f"✅ <b>Tasdiqlandi!</b> Admin: {escape_html(call.from_user.first_name)}"
            if call.message.content_type in ('photo', 'video', 'document', 'audio', 'voice'):
                new_caption = f"{call.message.caption or ''}\n\n{approve_text}"
                safe_execute(
                    bot.edit_message_caption,
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    caption=new_caption[:1024],
                    parse_mode='HTML'
                )
            else:
                safe_edit_message_text(
                    f"{call.message.text or ''}\n\n{approve_text}",
                    call.message.chat.id, call.message.message_id, parse_mode='HTML'
                )

            safe_send_message(
                submission['user_id'],
                f"🎉 Tabriklaymiz, {escape_html(submission['full_name'])}!\n\n"
                f"✅ Vazifangiz tasdiqlandi!\n"
                f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                f"🔢 ID: #{submission_id}",
                parse_mode='HTML'
            )
            safe_answer_callback_query(call.id, "✅ Tasdiqlandi!")

        # --- Rad etish ---
        elif data.startswith('reject_with_reason_'):
            user_states[user_id] = f'rejecting_reason_{submission_id}'
            safe_answer_callback_query(call.id, "✍️ Sabab yozing...")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('❌ Bekor qilish'))
            safe_send_message(
                call.message.chat.id,
                f"📝 <b>Rad etish sababini yozing:</b>\n\n"
                f"👤 O'quvchi: {escape_html(submission['full_name'])}\n"
                f"🔢 ID: #{submission_id}\n\n"
                f"Bu izoh o'quvchiga yuboriladi.",
                parse_mode='HTML',
                reply_markup=markup
            )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id].startswith('rejecting_reason_')
    )
    def save_rejection_reason(message):
        user_id = message.from_user.id
        submission_id = int(user_states[user_id].split('_')[-1])
        reason = message.text.strip()

        if reason == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if len(reason) < 5:
            safe_send_message(message.chat.id, "❌ Juda qisqa sabab. Batafsilroq yozing.")
            return

        safe_db_execute(
            'UPDATE submissions SET status = "rejected", reviewed_at = ?, reviewer_id = ?, rejection_reason = ? WHERE id = ?',
            (datetime.now(), user_id, reason, submission_id), commit=True
        )
        update_statistics('rejected')

        submission = safe_db_execute(
            'SELECT user_id, full_name FROM submissions WHERE id = ?', (submission_id,), fetch_one=True
        )
        if submission:
            safe_send_message(
                submission['user_id'],
                f"😔 Kechirasiz, {escape_html(submission['full_name'])}!\n\n"
                f"❌ Vazifangiz rad etildi.\n"
                f"📝 <b>Sabab:</b> {escape_html(reason)}\n\n"
                f"🔢 ID: #{submission_id}\n"
                f"🔄 Qayta topshirishingiz mumkin (max 3 marta)!",
                parse_mode='HTML'
            )

        clear_user_state(user_id)
        safe_send_message(
            message.chat.id,
            f"✅ Rad etildi!\n\n"
            f"👤 {escape_html(submission['full_name'] if submission else '?')}\n"
            f"🔢 ID: #{submission_id}",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )

    # ==================== BEKOR QILISH ====================
    @bot.message_handler(func=lambda m: m.text == '❌ Bekor qilish')
    def handle_cancel(message):
        user_id = message.from_user.id
        had_state = user_id in user_states
        clear_user_state(user_id)
        kb = get_admin_keyboard() if is_admin(user_id) else get_main_keyboard()
        msg = "❌ Operatsiya bekor qilindi." if had_state else "❌ Bekor."
        safe_send_message(message.chat.id, msg, reply_markup=kb)

    # ==================== IT MISOL: YARATISH (ADMIN) ====================
    @bot.message_handler(func=lambda m: m.text == '🏆 IT Misol' and is_admin(m.from_user.id))
    def start_contest_admin(message):
        user_id = message.from_user.id
        # Oldingi qolib ketgan state ni tozalash
        if user_id in user_states:
            logger.info(f"🧹 IT Misol boshlashdan oldin state tozalandi: {user_states[user_id]}")
            clear_user_state(user_id)
        user_states[user_id] = 'contest_step1_problem'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('❌ Bekor qilish'))
        safe_send_message(
            message.chat.id,
            "🏆 <b>IT Misol yaratish</b>\n\n"
            "<b>1-qadam:</b> Misol matnini kiriting:\n\n"
            "Masalan: <code>2 + 2 * 3 = ?</code>",
            parse_mode='HTML',
            reply_markup=markup
        )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id] == 'contest_step1_problem'
    )
    def contest_step1(message):
        user_id = message.from_user.id
        if message.text.strip() == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if len(message.text.strip()) < 3:
            safe_send_message(message.chat.id, "❌ Juda qisqa! Qayta kiriting.")
            return

        problem_text = message.text.strip()
        user_states[user_id] = {'step': 'step2_answer', 'problem': problem_text}
        safe_send_message(
            message.chat.id,
            f"✅ Misol saqlandi.\n\n"
            f"<b>2-qadam:</b> To'g'ri javobni kiriting:\n\n"
            f"Masalan: <code>8</code>",
            parse_mode='HTML'
        )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], dict)
            and user_states[m.from_user.id].get('step') == 'step2_answer'
        )
    )
    def contest_step2(message):
        user_id = message.from_user.id
        if message.text.strip() == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if not message.text.strip():
            safe_send_message(message.chat.id, "❌ Javob bo'sh bo'lishi mumkin emas!")
            return

        user_states[user_id]['correct_answer'] = message.text.strip()
        user_states[user_id]['step'] = 'step3_media'

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        markup.add(
            types.KeyboardButton('⏭ Mediasiz davom etish'),
            types.KeyboardButton('❌ Bekor qilish')
        )
        safe_send_message(
            message.chat.id,
            "✅ Javob saqlandi.\n\n"
            "<b>3-qadam:</b> Misol uchun rasm yoki video yuboring.\n"
            "(Ixtiyoriy — mediasiz ham davom etishingiz mumkin)",
            parse_mode='HTML',
            reply_markup=markup
        )

    # BUG FIX #3: Har bir content_type uchun alohida handler
    @bot.message_handler(
        content_types=['photo'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], dict)
            and user_states[m.from_user.id].get('step') == 'step3_media'
        )
    )
    def contest_step3_photo(message):
        user_id = message.from_user.id
        user_states[user_id]['media_file_id'] = message.photo[-1].file_id
        user_states[user_id]['media_type'] = 'photo'
        user_states[user_id]['step'] = 'step4_deadline'
        _ask_deadline(message.chat.id)

    @bot.message_handler(
        content_types=['video'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], dict)
            and user_states[m.from_user.id].get('step') == 'step3_media'
        )
    )
    def contest_step3_video(message):
        user_id = message.from_user.id
        user_states[user_id]['media_file_id'] = message.video.file_id
        user_states[user_id]['media_type'] = 'video'
        user_states[user_id]['step'] = 'step4_deadline'
        _ask_deadline(message.chat.id)

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], dict)
            and user_states[m.from_user.id].get('step') == 'step3_media'
        )
    )
    def contest_step3_text(message):
        user_id = message.from_user.id
        if message.text.strip() == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if message.text.strip() == '⏭ Mediasiz davom etish':
            user_states[user_id]['media_file_id'] = None
            user_states[user_id]['media_type'] = None
            user_states[user_id]['step'] = 'step4_deadline'
            _ask_deadline(message.chat.id)
        else:
            safe_send_message(message.chat.id, "❌ Faqat rasm, video yoki '⏭ Mediasiz davom etish' tugmasini bosing!")

    def _ask_deadline(chat_id):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
        markup.add(
            types.KeyboardButton('⏱ 5 daqiqa'),
            types.KeyboardButton('⏱ 10 daqiqa'),
            types.KeyboardButton('⏱ 15 daqiqa'),
            types.KeyboardButton('⏱ 30 daqiqa'),
            types.KeyboardButton('⏱ 1 soat'),
            types.KeyboardButton('❌ Bekor qilish')
        )
        safe_send_message(
            chat_id,
            "✅ Media saqlandi!\n\n<b>4-qadam:</b> Muddat tanlang:",
            parse_mode='HTML',
            reply_markup=markup
        )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], dict)
            and user_states[m.from_user.id].get('step') == 'step4_deadline'
        )
    )
    def contest_step4_deadline(message):
        global active_contest
        user_id = message.from_user.id

        if message.text.strip() == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return

        minutes_map = {
            '5 daqiqa': 5, '10 daqiqa': 10,
            '15 daqiqa': 15, '30 daqiqa': 30, '1 soat': 60
        }
        minutes = None
        for key, val in minutes_map.items():
            if key in message.text:
                minutes = val
                break

        if not minutes:
            safe_send_message(message.chat.id, "❌ Ro'yxatdan muddat tanlang!")
            return

        state = user_states[user_id]
        problem_text = state['problem']
        correct_answer = state['correct_answer']
        media_file_id = state.get('media_file_id')
        media_type = state.get('media_type')

        deadline = datetime.now() + timedelta(minutes=minutes)

        contest_id = safe_db_execute(
            'INSERT INTO contests (problem_text, correct_answer, deadline, media_file_id, media_type) VALUES (?, ?, ?, ?, ?)',
            (problem_text, correct_answer, deadline, media_file_id, media_type),
            commit=True
        )
        if not contest_id:
            safe_send_message(message.chat.id, "❌ Xatolik! Qayta urinib ko'ring.")
            return

        active_contest = contest_id

        # BUG FIX #4: O'quvchilarga to'g'ri javob CHIQMASIN
        # Kanalga ham to'g'ri javob chiqmasin!
        student_caption = (
            f"🏆 <b>YANGI IT MISOL!</b>\n\n"
            f"❓ <b>Misol:</b>\n{escape_html(problem_text)}\n\n"
            f"⏱ Muddat: {minutes} daqiqa ({deadline.strftime('%H:%M')} gacha)\n"
            f"🏁 Birinchi to'g'ri javob g'olib!\n\n"
            f"✍️ <b>'Javob yuborish'</b> tugmasini bosing!\n"
            f"🔢 Contest ID: #{contest_id}"
        )

        student_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        student_markup.add(types.KeyboardButton('✍️ Javob yuborish'))

        students = get_all_students()
        sent_count = 0
        for student in students:
            if not is_admin(student['user_id']):
                try:
                    if media_file_id and media_type == 'photo':
                        safe_send_photo(
                            student['user_id'], media_file_id,
                            caption=student_caption, parse_mode='HTML',
                            reply_markup=student_markup
                        )
                    elif media_file_id and media_type == 'video':
                        safe_execute(
                            bot.send_video, student['user_id'], media_file_id,
                            caption=student_caption, parse_mode='HTML',
                            reply_markup=student_markup
                        )
                    else:
                        safe_send_message(
                            student['user_id'], student_caption,
                            parse_mode='HTML', reply_markup=student_markup
                        )
                    sent_count += 1
                except Exception as e:
                    logger.error(f"Contest yuborishda xato ({student['user_id']}): {e}")

        # Kanalga ham to'g'ri javobsiz yuborish
        if media_file_id and media_type == 'photo':
            safe_send_photo(CHANNEL_ID, media_file_id, caption=student_caption, parse_mode='HTML')
        elif media_file_id and media_type == 'video':
            safe_execute(bot.send_video, CHANNEL_ID, media_file_id, caption=student_caption, parse_mode='HTML')
        else:
            safe_send_message(CHANNEL_ID, student_caption, parse_mode='HTML')

        clear_user_state(user_id)
        media_label = 'Rasm ✅' if media_type == 'photo' else 'Video ✅' if media_type == 'video' else "Yo'q"
        safe_send_message(
            message.chat.id,
            f"✅ <b>Contest boshlandi!</b>\n\n"
            f"🏆 ID: #{contest_id}\n"
            f"❓ Misol: {escape_html(problem_text[:60])}\n"
            f"🔑 To'g'ri javob: <b>{escape_html(correct_answer)}</b>\n"
            f"⏱ Muddat: {minutes} daqiqa\n"
            f"📎 Media: {media_label}\n"
            f"📢 Yuborildi: {sent_count} ta o'quvchi",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
        logger.info(f"✅ Contest yaratildi: #{contest_id}, javob: {correct_answer}")

    # ==================== CONTEST JAVOB (O'QUVCHI - FAQAT MATN) ====================
    @bot.message_handler(func=lambda m: m.text == '✍️ Javob yuborish')
    def submit_contest_start(message):
        global active_contest
        user_id = message.from_user.id
        # Oldingi state ni tozalash
        if user_id in user_states and not str(user_states.get(user_id, '')).startswith('contest_answer_'):
            logger.info(f"🧹 Contest javobdan oldin state tozalandi: {user_states[user_id]}")
            clear_user_state(user_id)

        if not is_registered(user_id):
            safe_send_message(message.chat.id, "❌ Avval ro'yxatdan o'ting!")
            return
        if not active_contest:
            safe_send_message(message.chat.id, "❌ Faol musobaqa yo'q!")
            return

        contest = safe_db_execute(
            'SELECT * FROM contests WHERE id = ? AND is_active = 1', (active_contest,), fetch_one=True
        )
        if not contest:
            safe_send_message(message.chat.id, "❌ Musobaqa yakunlangan!")
            return

        deadline = contest['deadline']
        if isinstance(deadline, str):
            deadline = datetime.fromisoformat(deadline)
        if datetime.now() > deadline:
            safe_send_message(message.chat.id, "⏰ Vaqt tugagan!")
            return

        user_states[user_id] = f'contest_answer_{active_contest}'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('❌ Bekor qilish'))
        mins_left = max(0, int((deadline - datetime.now()).total_seconds() / 60))
        safe_send_message(
            message.chat.id,
            f"✍️ <b>Javobingizni yozing:</b>\n\n"
            f"❓ {escape_html(contest['problem_text'])}\n\n"
            f"⏱ Qolgan vaqt: ~{mins_left} daqiqa\n\n"
            f"📝 Faqat <b>matn</b> (javob) yuboring:",
            parse_mode='HTML',
            reply_markup=markup
        )

    # O'quvchi javob beradi — FAQAT MATN qabul qilinadi
    @bot.message_handler(
        content_types=['text'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], str)
            and user_states[m.from_user.id].startswith('contest_answer_')
        )
    )
    def receive_contest_answer(message):
        user_id = message.from_user.id
        contest_id = int(user_states[user_id].replace('contest_answer_', ''))
        answer = message.text.strip()

        if answer == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_main_keyboard())
            return

        student = get_student_info(user_id)
        contest = safe_db_execute(
            'SELECT * FROM contests WHERE id = ?', (contest_id,), fetch_one=True
        )
        if not contest:
            safe_send_message(message.chat.id, "❌ Contest topilmadi!")
            return

        deadline = contest['deadline']
        if isinstance(deadline, str):
            deadline = datetime.fromisoformat(deadline)
        if datetime.now() > deadline:
            safe_send_message(message.chat.id, "⏰ Vaqt tugagan!")
            clear_user_state(user_id)
            return

        correct = contest['correct_answer'].strip().lower()
        is_correct = (
            answer.lower() == correct or
            SequenceMatcher(None, answer.lower(), correct).ratio() > 0.85
        )

        rank_position = None
        if is_correct:
            rank_row = safe_db_execute(
                'SELECT COUNT(*) as c FROM contest_submissions WHERE contest_id = ? AND is_correct = 1',
                (contest_id,), fetch_one=True
            )
            rank_position = (rank_row['c'] if rank_row else 0) + 1

        safe_db_execute(
            'INSERT INTO contest_submissions (contest_id, user_id, full_name, answer, is_correct, rank_position, submitted_at) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (contest_id, user_id, student['full_name'], answer, 1 if is_correct else 0, rank_position, datetime.now()),
            commit=True
        )
        clear_user_state(user_id)

        if is_correct:
            emoji = "🥇" if rank_position == 1 else "🥈" if rank_position == 2 else "🥉" if rank_position == 3 else "🏅"
            safe_send_message(
                message.chat.id,
                f"🎉 <b>TABRIKLAYMIZ!</b>\n\n"
                f"✅ To'g'ri javob!\n"
                f"{emoji} Sizning o'rningiz: <b>{rank_position}</b>\n"
                f"⏰ Vaqt: {datetime.now().strftime('%H:%M:%S')}",
                parse_mode='HTML',
                reply_markup=get_main_keyboard()
            )
            # Kanalga: to'g'ri javobni chiqarmaslik (faqat kim yechdi)
            safe_send_message(
                CHANNEL_ID,
                f"🏆 <b>Contest #{contest_id}</b>\n\n"
                f"{emoji} {rank_position}-o'rin: <b>{escape_html(student['full_name'])}</b>\n"
                f"⏰ {datetime.now().strftime('%H:%M:%S')}",
                parse_mode='HTML'
            )
        else:
            mins_left = max(0, int((deadline - datetime.now()).total_seconds() / 60))
            safe_send_message(
                message.chat.id,
                f"❌ Noto'g'ri javob!\n\n"
                f"💡 Qayta urinib ko'ring!\n"
                f"⏱ Qolgan: ~{mins_left} daqiqa",
                reply_markup=get_main_keyboard()
            )

    # ==================== REYTING ====================
    @bot.message_handler(func=lambda m: m.text == '🏆 Reyting')
    def show_leaderboard(message):
        global active_contest
        if not active_contest:
            safe_send_message(message.chat.id, "❌ Faol musobaqa yo'q!")
            return

        results = safe_db_execute(
            'SELECT full_name, submitted_at, rank_position FROM contest_submissions '
            'WHERE contest_id = ? AND is_correct = 1 ORDER BY rank_position ASC',
            (active_contest,), fetch_all=True
        )
        contest = safe_db_execute(
            'SELECT problem_text, deadline FROM contests WHERE id = ?', (active_contest,), fetch_one=True
        )

        deadline = contest['deadline'] if contest else None
        if isinstance(deadline, str):
            try:
                deadline = datetime.fromisoformat(deadline)
            except:
                deadline = None

        text = f"🏆 <b>Reyting (Contest #{active_contest})</b>\n\n"
        if contest:
            text += f"❓ {escape_html(contest['problem_text'][:60])}\n"
            text += f"⏱ {deadline.strftime('%H:%M') if deadline else '?'}\n\n"

        if not results:
            text += "❌ Hozircha to'g'ri javob yo'q!"
        else:
            for res in results:
                r = res['rank_position']
                emoji = "🥇" if r == 1 else "🥈" if r == 2 else "🥉" if r == 3 else "🏅"
                t = res['submitted_at']
                t_str = t.strftime('%H:%M:%S') if t else '?'
                text += f"{emoji} {r}-o'rin: {escape_html(res['full_name'])} — {t_str}\n"

        safe_send_message(message.chat.id, text, parse_mode='HTML')

    # ==================== ADMIN BOSHQARUV ====================
    @bot.message_handler(func=lambda m: m.text == '➕ Admin boshqaruv' and is_admin(m.from_user.id))
    def admin_management(message):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("➕ Yangi admin qo'shish", callback_data='admin_add'),
            types.InlineKeyboardButton("📋 Adminlar ro'yxati", callback_data='admin_list'),
            types.InlineKeyboardButton("❌ Admin o'chirish", callback_data='admin_remove_list')
        )
        safe_send_message(
            message.chat.id,
            "👨‍💼 <b>Admin boshqaruvi</b>",
            parse_mode='HTML',
            reply_markup=markup
        )

    @bot.callback_query_handler(
        func=lambda call: call.data in ('admin_add', 'admin_list', 'admin_remove_list', 'admin_back')
    )
    def admin_mgmt_callback(call):
        user_id = call.from_user.id
        if not is_admin(user_id):
            safe_answer_callback_query(call.id, "❌ Ruxsat yo'q!")
            return

        if call.data == 'admin_add':
            user_states[user_id] = 'adding_admin_id'
            safe_answer_callback_query(call.id)
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('❌ Bekor qilish'))
            safe_send_message(
                call.message.chat.id,
                "➕ <b>Yangi admin qo'shish</b>\n\n"
                "Foydalanuvchining Telegram <b>ID</b> raqamini kiriting:\n\n"
                "💡 ID ni bilish uchun foydalanuvchi @userinfobot ga yozsin.",
                parse_mode='HTML',
                reply_markup=markup
            )

        elif call.data == 'admin_list':
            admins = get_all_admins()
            text = "📋 <b>Adminlar ro'yxati</b>\n\n"
            text += "🔒 <b>Doimiy (kod ichida):</b>\n"
            for aid in ADMIN_IDS:
                s = get_student_info(aid)
                name = s['full_name'] if s else f"ID:{aid}"
                text += f"  • {escape_html(name)} (<code>{aid}</code>)\n"
            text += "\n📝 <b>Qo'shilgan (DB):</b>\n"
            if admins:
                for a in admins:
                    un = f"@{a['username']}" if a['username'] else "—"
                    added = a['added_at'].strftime('%d.%m.%Y') if a['added_at'] else '?'
                    text += f"  • {escape_html(a['full_name'])} ({un})\n"
                    text += f"    <code>{a['user_id']}</code> | {added}\n"
            else:
                text += "  Hech kim yo'q\n"
            safe_edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode='HTML')
            safe_answer_callback_query(call.id, "✅ Yuklandi")

        elif call.data == 'admin_remove_list':
            admins = get_all_admins()
            if not admins:
                safe_answer_callback_query(call.id, "❌ O'chiriladigan admin yo'q!", show_alert=True)
                return
            markup = types.InlineKeyboardMarkup(row_width=1)
            for a in admins:
                markup.add(types.InlineKeyboardButton(
                    f"❌ {a['full_name']} ({a['user_id']})",
                    callback_data=f"admin_rm_{a['user_id']}"
                ))
            markup.add(types.InlineKeyboardButton("🔙 Orqaga", callback_data='admin_back'))
            safe_edit_message_text(
                "❌ <b>Qaysi adminni o'chirish?</b>",
                call.message.chat.id, call.message.message_id,
                parse_mode='HTML', reply_markup=markup
            )
            safe_answer_callback_query(call.id)

        elif call.data == 'admin_back':
            markup = types.InlineKeyboardMarkup(row_width=1)
            markup.add(
                types.InlineKeyboardButton("➕ Yangi admin qo'shish", callback_data='admin_add'),
                types.InlineKeyboardButton("📋 Adminlar ro'yxati", callback_data='admin_list'),
                types.InlineKeyboardButton("❌ Admin o'chirish", callback_data='admin_remove_list')
            )
            safe_edit_message_text(
                "👨‍💼 <b>Admin boshqaruvi</b>",
                call.message.chat.id, call.message.message_id,
                parse_mode='HTML', reply_markup=markup
            )
            safe_answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('admin_rm_'))
    def admin_remove_callback(call):
        user_id = call.from_user.id
        if not is_admin(user_id):
            safe_answer_callback_query(call.id, "❌ Ruxsat yo'q!")
            return

        target_id = int(call.data.replace('admin_rm_', ''))
        if target_id in ADMIN_IDS:
            safe_answer_callback_query(call.id, "❌ Doimiy adminni o'chirib bo'lmaydi!", show_alert=True)
            return

        a = safe_db_execute('SELECT full_name FROM admins WHERE user_id = ?', (target_id,), fetch_one=True)
        name = a['full_name'] if a else str(target_id)

        safe_db_execute('UPDATE admins SET is_active = 0 WHERE user_id = ?', (target_id,), commit=True)
        safe_edit_message_text(
            f"✅ <b>{escape_html(name)}</b> admin lavozimidan olindi.",
            call.message.chat.id, call.message.message_id, parse_mode='HTML'
        )
        safe_answer_callback_query(call.id, "✅ O'chirildi")
        safe_send_message(target_id, "⚠️ Sizning admin huquqlaringiz bekor qilindi.", reply_markup=get_main_keyboard())
        logger.info(f"Admin o'chirildi: {target_id} ({name})")

    # Admin ID olish
    @bot.message_handler(
        content_types=['text'],
        func=lambda m: m.from_user.id in user_states and user_states[m.from_user.id] == 'adding_admin_id'
    )
    def add_admin_receive_id(message):
        user_id = message.from_user.id
        text = message.text.strip()

        if text == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return

        try:
            new_id = int(text)
        except ValueError:
            safe_send_message(message.chat.id, "❌ Noto'g'ri format! Faqat raqam kiriting.")
            return

        if new_id in ADMIN_IDS:
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "✅ Bu foydalanuvchi doimiy admin!", reply_markup=get_admin_keyboard())
            return

        existing = safe_db_execute('SELECT is_active FROM admins WHERE user_id = ?', (new_id,), fetch_one=True)
        if existing and existing['is_active'] == 1:
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "✅ Bu foydalanuvchi allaqachon admin!", reply_markup=get_admin_keyboard())
            return

        student = get_student_info(new_id)
        if student:
            # Ro'yxatdan o'tgan — to'g'ridan admin qilish
            save_admin_to_db(new_id, student['full_name'], student.get('username'), user_id)
            clear_user_state(user_id)
            safe_send_message(
                message.chat.id,
                f"✅ <b>{escape_html(student['full_name'])}</b> admin qilindi!\n🆔 {new_id}",
                parse_mode='HTML', reply_markup=get_admin_keyboard()
            )
            safe_send_message(
                new_id,
                "🎉 Siz admin sifatida qo'shildingiz!\n/start buyrug'ini yuboring."
            )
        else:
            # Ro'yxatdan o'tmagan — ismini so'rash
            user_states[user_id] = f'adding_admin_name_{new_id}'
            safe_send_message(
                message.chat.id,
                f"✅ ID: <code>{new_id}</code>\n\n"
                "Bu foydalanuvchi ro'yxatdan o'tmagan.\n"
                "Uning to'liq ismini kiriting:",
                parse_mode='HTML'
            )

    @bot.message_handler(
        content_types=['text'],
        func=lambda m: (
            m.from_user.id in user_states
            and isinstance(user_states[m.from_user.id], str)
            and user_states[m.from_user.id].startswith('adding_admin_name_')
        )
    )
    def add_admin_receive_name(message):
        user_id = message.from_user.id
        full_name = message.text.strip()
        new_id = int(user_states[user_id].replace('adding_admin_name_', ''))

        if full_name == '❌ Bekor qilish':
            clear_user_state(user_id)
            safe_send_message(message.chat.id, "❌ Bekor qilindi.", reply_markup=get_admin_keyboard())
            return
        if len(full_name) < 3:
            safe_send_message(message.chat.id, "❌ Juda qisqa ism!")
            return

        save_admin_to_db(new_id, full_name, None, user_id)
        clear_user_state(user_id)
        safe_send_message(
            message.chat.id,
            f"✅ <b>{escape_html(full_name)}</b> admin qilindi!\n🆔 {new_id}",
            parse_mode='HTML', reply_markup=get_admin_keyboard()
        )
        safe_send_message(
            new_id,
            "🎉 Siz admin sifatida qo'shildingiz!\n/start buyrug'ini yuboring."
        )

    # ==================== EXCEL EXPORT ====================
    @bot.message_handler(func=lambda m: m.text == '📥 Excel' and is_admin(m.from_user.id))
    def export_excel(message):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            submissions = safe_db_execute(
                '''SELECT s.id, s.full_name, s.homework_text, s.submitted_at, s.status,
                   s.rejection_reason, a.homework_text as assignment_text, a.assignment_date
                   FROM submissions s
                   LEFT JOIN assignments a ON s.assignment_id = a.id
                   ORDER BY s.submitted_at DESC''',
                fetch_all=True
            )
            if not submissions:
                safe_send_message(message.chat.id, "❌ Hech qanday topshiriq yo'q!")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Topshiriqlar"
            headers = ['ID', 'Ism', 'Vazifa matni', 'Sana', 'Holat', 'Rad sababi', 'Berilgan vazifa', 'Sana']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            for sub in submissions:
                holat = '✅ Tasdiqlangan' if sub['status'] == 'approved' else \
                        '❌ Rad etilgan' if sub['status'] == 'rejected' else '⏳ Kutilmoqda'
                ws.append([
                    sub['id'],
                    sub['full_name'],
                    (sub['homework_text'] or '')[:200],
                    sub['submitted_at'].strftime('%d.%m.%Y %H:%M') if sub['submitted_at'] else '',
                    holat,
                    sub['rejection_reason'] or '',
                    (sub['assignment_text'] or '')[:100],
                    str(sub['assignment_date']) if sub['assignment_date'] else ''
                ])

            for col, width in zip(['A','B','C','D','E','F','G','H'], [8,25,40,18,18,30,40,15]):
                ws.column_dimensions[col].width = width

            filename = f"topshiriqlar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            with open(filename, 'rb') as f:
                safe_send_document(
                    message.chat.id, f,
                    caption=f"📊 Barcha topshiriqlar ({len(submissions)} ta)"
                )
            os.remove(filename)

        except ImportError:
            safe_send_message(message.chat.id, "❌ openpyxl o'rnatilmagan!\n\n<code>pip install openpyxl</code>", parse_mode='HTML')
        except Exception as e:
            logger.error(f"Excel export: {e}")
            safe_send_message(message.chat.id, f"❌ Xatolik: {e}")

    # ==================== ADMIN PANEL ====================
    @bot.message_handler(func=lambda m: m.text == '👨‍💼 Admin panel' and is_admin(m.from_user.id))
    def admin_panel(message):
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("👥 O'quvchilar ro'yxati", callback_data='admin_students'),
            types.InlineKeyboardButton("📊 Umumiy statistika", callback_data='admin_stats'),
            types.InlineKeyboardButton("🗑️ Barchasini tozalash", callback_data='admin_clear')
        )
        safe_send_message(
            message.chat.id,
            "👨‍💼 <b>Admin Panel</b>",
            parse_mode='HTML',
            reply_markup=markup
        )

    @bot.callback_query_handler(func=lambda call: call.data in ('admin_students', 'admin_stats', 'admin_clear', 'confirm_clear', 'cancel_clear'))
    def admin_panel_callbacks(call):
        user_id = call.from_user.id
        if not is_admin(user_id):
            safe_answer_callback_query(call.id, "❌ Ruxsat yo'q!")
            return

        if call.data == 'admin_students':
            students = get_all_students()
            text = "👥 <b>O'quvchilar:</b>\n\n" if students else "👥 Ro'yxat bo'sh."
            for i, s in enumerate(students, 1):
                un = f"@{s['username']}" if s['username'] else "—"
                reg = s['registered_at'].strftime('%d.%m.%Y') if s['registered_at'] else '?'
                text += f"{i}. {escape_html(s['full_name'])} ({un})\n   📅 {reg}\n\n"
            # 4096 limit
            if len(text) > 4000:
                text = text[:4000] + "\n...(davomi bor)"
            safe_edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode='HTML')
            safe_answer_callback_query(call.id, "✅ Yuklandi")

        elif call.data == 'admin_stats':
            overall = safe_db_execute(
                'SELECT SUM(total_submissions) as t, SUM(approved_submissions) as a, SUM(rejected_submissions) as r FROM statistics',
                fetch_one=True
            )
            today = datetime.now().strftime('%Y-%m-%d')
            today_s = safe_db_execute(
                'SELECT * FROM statistics WHERE date = ?', (today,), fetch_one=True
            )
            text = (
                f"📊 <b>Umumiy statistika</b>\n\n"
                f"📈 Jami: {overall['t'] or 0}\n"
                f"✅ Tasdiqlangan: {overall['a'] or 0}\n"
                f"❌ Rad etilgan: {overall['r'] or 0}\n\n"
                f"📅 <b>Bugun:</b>\n"
            )
            if today_s:
                text += (
                    f"• Topshirilgan: {today_s['total_submissions'] or 0}\n"
                    f"• Tasdiqlangan: {today_s['approved_submissions'] or 0}\n"
                    f"• Rad etilgan: {today_s['rejected_submissions'] or 0}\n"
                )
            else:
                text += "• Hech narsa yo'q\n"
            safe_edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode='HTML')
            safe_answer_callback_query(call.id, "✅ Yuklandi")

        elif call.data == 'admin_clear':
            markup = types.InlineKeyboardMarkup()
            markup.add(
                types.InlineKeyboardButton("✅ Ha, tozalash", callback_data='confirm_clear'),
                types.InlineKeyboardButton("❌ Bekor", callback_data='cancel_clear')
            )
            safe_edit_message_text(
                "⚠️ <b>Diqqat!</b>\n\nBarcha topshiriqlar, statistika, contestlar o'chiriladi.\n"
                "Adminlar va o'quvchilar saqlanadi.\n\nTasdiqlaysizmi?",
                call.message.chat.id, call.message.message_id,
                parse_mode='HTML', reply_markup=markup
            )

        elif call.data == 'confirm_clear':
            try:
                conn = get_db_connection()
                if conn:
                    cur = conn.cursor()
                    cur.execute('DELETE FROM submissions')
                    cur.execute('DELETE FROM statistics')
                    cur.execute('DELETE FROM contest_submissions')
                    cur.execute('DELETE FROM contests')
                    cur.execute('DELETE FROM assignments')
                    conn.commit()
                    conn.close()
                    global active_contest
                    active_contest = None
                    text = "✅ Barcha ma'lumotlar tozalandi!\nO'quvchilar va adminlar saqlanib qoldi."
                else:
                    text = "❌ Xatolik!"
            except Exception as e:
                logger.error(f"Clear error: {e}")
                text = f"❌ Xatolik: {e}"
            safe_edit_message_text(text, call.message.chat.id, call.message.message_id, parse_mode='HTML')
            safe_answer_callback_query(call.id, "✅ Bajarildi")

        elif call.data == 'cancel_clear':
            safe_edit_message_text(
                "✅ Tozalash bekor qilindi.",
                call.message.chat.id, call.message.message_id, parse_mode='HTML'
            )
            safe_answer_callback_query(call.id, "Bekor qilindi")

    logger.info("✅ Barcha handlerlar ro'yxatdan o'tdi")

# ==================== ISHGA TUSHIRISH ====================
def start_bot():
    global bot
    logger.info("🚀 Bot ishga tushmoqda...")

    if not init_db():
        logger.critical("❌ Database yaratilmadi!")
        return False

    bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)

    try:
        bot.remove_webhook()
        time.sleep(1)
        bot.get_updates(offset=-1)
        logger.info("✅ Webhook tozalandi")
    except Exception as e:
        logger.warning(f"⚠️ Webhook tozalashda xato (normal): {e}")

    register_handlers(bot)
    logger.info("✅ Bot tayyor. Polling boshlandi...")

    retry_count = 0
    max_retries = 5

    while retry_count < max_retries:
        try:
            bot.infinity_polling(
                timeout=30,
                long_polling_timeout=60,
                allowed_updates=["message", "callback_query"],
                skip_pending=False,
                none_stop=True
            )
            break
        except telebot.apihelper.ApiTelegramException as e:
            if "409" in str(e) or "Conflict" in str(e):
                logger.error("❌ 409 Conflict: Boshqa bot instance ishlayapti!")
                time.sleep(10)
                retry_count += 1
                try:
                    bot.remove_webhook()
                    time.sleep(2)
                    bot.get_updates(offset=-1)
                except:
                    pass
                if retry_count >= max_retries:
                    logger.critical("❌ Bot ishlamadi. Eski instanceni to'xtating!")
                    return False
            else:
                raise
        except KeyboardInterrupt:
            logger.info("⏹ Bot to'xtatildi")
            break
        except Exception as e:
            logger.error(f"❌ Polling xato: {e}")
            retry_count += 1
            if retry_count >= max_retries:
                return False
            time.sleep(5 * retry_count)
            bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)
            register_handlers(bot)

    return True

if __name__ == '__main__':
    try:
        start_bot()
    except Exception as e:
        logger.critical(f"❌ Fatal: {e}")
        logger.critical(traceback.format_exc())
        sys.exit(1)
